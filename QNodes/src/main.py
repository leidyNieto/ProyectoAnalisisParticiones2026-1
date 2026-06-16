"""
Punto de entrada de QNodes en modo lote sobre DatosPruebas2026_1.xlsx.

Lee el Excel de entrada (compartido con GeoMIP, en GeoMIP/results/), ejecuta
QNodes con k-particiones (k = 2 … 5) para cada prueba de cada hoja y escribe
los resultados en las columnas QNodes del Excel de resultados. Si el Excel de
resultados no existe, lo crea copiando la plantilla de entrada.

Control por variables de entorno (igual estilo que GeoMIP):
    QNODES_INPUT_XLSX   ruta del Excel de entrada
    QNODES_OUTPUT_XLSX  ruta del Excel de resultados
    QNODES_HOJA         procesa solo las hojas cuyo nombre contenga este texto
    QNODES_CANTIDAD     limita la cantidad de pruebas por hoja
    QNODES_TIMEOUT      segundos máximos por prueba (default 3600)
"""

import multiprocessing
import os
import shutil
import traceback
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from src.controllers.manager import Manager
from src.funcs.excel_pruebas2026 import (
    escribir_resultado_fila,
    iter_hojas_pruebas,
    parsear_nombre_hoja,
)
from src.strategies.q_nodes import QNodes

QNODES_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = Path(__file__).resolve().parents[2]
GEOMIP_ROOT = PROJECT_ROOT / "GeoMIP"

def _resolver_excel_entrada(nombre: str = "DatosPruebas2026_1.xlsx") -> Path:
    """Localiza el Excel de pruebas admitiendo varias ubicaciones de 'results'."""
    candidatos = (
        PROJECT_ROOT / "results" / nombre,
        GEOMIP_ROOT / "results" / nombre,
        QNODES_ROOT / "results" / nombre,
    )
    for ruta in candidatos:
        if ruta.exists():
            return ruta
    return candidatos[0]


# El Excel de entrada/salida es compartido con GeoMIP (mismas redes => comparacion valida).
DEFAULT_EXCEL = _resolver_excel_entrada()


def resolver_tpm_path(estado_inicio: str, pagina: str = "A") -> Path:
    """Busca la TPM N{n}{pagina}.csv en las rutas conocidas."""
    n = len(estado_inicio)
    candidatos_letra = [pagina] + [c for c in "ABCDE" if c != pagina]
    # Se priorizan las TPMs compartidas de GeoMIP para que ambos frameworks
    # analicen exactamente la misma red (comparación válida).
    bases = (
        GEOMIP_ROOT / "data" / "samples",
        QNODES_ROOT / "src" / ".samples",
        QNODES_ROOT / ".samples",
    )
    for letra in candidatos_letra:
        sample_name = f"N{n}{letra}.csv"
        for base in bases:
            path = base / sample_name
            if path.exists():
                return path
    disponibles = []
    for base in bases:
        if base.exists():
            disponibles.extend(p.name for p in base.glob(f"N{n}*.csv"))
    extra = (
        f" Disponibles para N{n}: "
        f"{', '.join(sorted(set(disponibles))) or 'ninguno'}."
    )
    raise FileNotFoundError(f"No se encontro TPM para N{n} (pagina {pagina}).{extra}")


def asegurar_tpm(n_nodos: int, pagina: str) -> Path:
    """Devuelve la ruta de la TPM; si no existe la genera con Manager."""
    estado = "1" + "0" * (n_nodos - 1)
    auto_gen = os.getenv(
        "QNODES_AUTO_TPM", os.getenv("GEOMIP_AUTO_TPM", "1")
    ).lower() in ("1", "true", "si", "s")

    try:
        return resolver_tpm_path(estado, pagina)
    except FileNotFoundError:
        pass

    print(
        f"\n>>> TPM N{n_nodos}{pagina} no encontrada. "
        "Generando con Manager.generar_red ..."
    )
    gestor = Manager(estado_inicial=estado)
    nombre = gestor.generar_red(
        dimensiones=n_nodos,
        pagina=pagina,
        datos_discretos=False,
        auto_confirm=auto_gen,
    )
    if not nombre:
        raise FileNotFoundError(
            f"No se pudo generar N{n_nodos}{pagina}.csv. "
            "Para redes >1 GB use QNODES_AUTO_TPM=1."
        )
    path = gestor.ruta_base / nombre
    if not path.exists():
        raise FileNotFoundError(
            f"Generacion reporto {nombre} pero no existe en {path}"
        )
    print(f">>> TPM generada: {path}\n")
    return path


def _cargar_tpm(tpm_path: Path, nombre_hoja: str) -> np.ndarray | None:
    try:
        return np.genfromtxt(tpm_path, delimiter=",")
    except MemoryError:
        print(f"\n=== TPM demasiado grande para numpy: {tpm_path} ===")
        print("Intentando fallback con pandas (float32).")
        try:
            dtype = pd.Float32Dtype() if hasattr(pd, "Float32Dtype") else "float32"
            return pd.read_csv(tpm_path, header=None, dtype=dtype).values
        except Exception as exc:
            print(f"Error al cargar TPM con pandas: {exc}")
            print(f"Omitiendo hoja {nombre_hoja}.")
            return None


def _resolver_prueba(estado, condiciones, alcance, mecanismo, tpm) -> dict:
    analizador = QNodes(tpm)
    # mostrar_tabla=True imprime la tabla coloreada por-k de esta prueba.
    sol = analizador.aplicar_estrategia(
        estado, condiciones, alcance, mecanismo, mostrar_tabla=True
    )
    return {
        "particion": sol.particion,
        "perdida": sol.perdida,
        "tiempo": sol.tiempo_ejecucion,
        "k": sol.k_particiones,
        "por_k": sol.resultados_por_k or {},
    }


def _worker_prueba(args, queue):
    try:
        queue.put(_resolver_prueba(*args))
    except Exception:
        traceback.print_exc()
        queue.put(None)


def _fmt_num(val):
    if val is None:
        return None
    return str(val).replace(".", ",")


def _fila_resumen(resultado: dict) -> dict:
    por_k = resultado.get("por_k") or {}
    fila = {
        "Mejor k global": resultado.get("k"),
        "Mejor particion global": resultado.get("particion"),
        "Mejor perdida global": _fmt_num(resultado.get("perdida")),
        "Tiempo total (s)": _fmt_num(resultado.get("tiempo")),
    }
    for k in range(2, 6):
        r = por_k.get(k, {})
        fila[f"k{k} Particion"] = r.get("particion")
        fila[f"k{k} Perdida"] = _fmt_num(r.get("perdida"))
        fila[f"k{k} Tiempo (s)"] = _fmt_num(r.get("tiempo"))
    return fila


def ejecutar_datos_pruebas2026(
    ruta_entrada: Path,
    ruta_salida: Path | None = None,
    cantidad: int | None = None,
    hoja_filtro: str | None = None,
):
    """
    Procesa todas las hojas (o solo `hoja_filtro`) del Excel de entrada con
    QNodes y escribe en las columnas QNodes del Excel de resultados.
    """
    ruta_salida = ruta_salida or ruta_entrada.with_name(
        ruta_entrada.stem + "_resultados.xlsx"
    )
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    # Crear el Excel de resultados desde la plantilla si aún no existe.
    if not ruta_salida.exists():
        if not ruta_entrada.exists():
            raise FileNotFoundError(
                f"No existe el Excel de entrada ni el de resultados: {ruta_entrada}"
            )
        shutil.copy2(ruta_entrada, ruta_salida)
        print(f"Archivo de resultados creado desde plantilla: {ruta_salida}")

    wb = load_workbook(ruta_salida)
    resumen_filas = []
    timeout = int(os.getenv("QNODES_TIMEOUT", os.getenv("GEOMIP_TIMEOUT", "3600")))

    for nombre_hoja, pruebas in iter_hojas_pruebas(ruta_entrada):
        if hoja_filtro and hoja_filtro.lower() not in nombre_hoja.lower():
            continue
        if nombre_hoja not in wb.sheetnames:
            continue

        parsed = parsear_nombre_hoja(nombre_hoja)
        if not parsed:
            continue
        n_nodos, pagina = parsed

        try:
            tpm_path = asegurar_tpm(n_nodos, pagina)
        except FileNotFoundError as exc:
            print(f"\n=== OMITIDA hoja {nombre_hoja}: {exc} ===")
            continue

        tpm = _cargar_tpm(tpm_path, nombre_hoja)
        if tpm is None:
            continue

        ws = wb[nombre_hoja]

        # Encabezado de la hoja (una sola vez al empezar la hoja).
        print(f"\n=== Hoja: {nombre_hoja} | TPM: {tpm_path.name} ===")

        limite = cantidad if cantidad else len(pruebas)
        for prueba in pruebas[:limite]:
            print(
                f"  Prueba #{prueba.numero} | alcance={prueba.alcance} "
                f"mecanismo={prueba.mecanismo}"
            )

            queue = multiprocessing.Queue()
            args = (
                prueba.estado_inicial,
                prueba.condiciones,
                prueba.alcance,
                prueba.mecanismo,
                tpm,
            )
            proc = multiprocessing.Process(target=_worker_prueba, args=(args, queue))
            proc.start()
            proc.join(timeout=timeout)

            if proc.is_alive():
                proc.terminate()
                proc.join()
                resultado = None
                print(f"  Prueba #{prueba.numero} - TIMEOUT")
            else:
                resultado = queue.get() if not queue.empty() else None

            if resultado:
                escribir_resultado_fila(ws, prueba.fila_excel, resultado["por_k"])
                print(
                    f"  -> Mejor k={resultado['k']} "
                    f"perdida={resultado['perdida']:.4f} "
                    f"tiempo={resultado['tiempo']:.3f}s  (escrito en Excel)"
                )
                resumen_filas.append(
                    {
                        "Hoja": nombre_hoja,
                        "Prueba": prueba.numero,
                        "Alcance": prueba.alcance,
                        "Mecanismo": prueba.mecanismo,
                        **_fila_resumen(resultado),
                    }
                )
            else:
                resumen_filas.append(
                    {
                        "Hoja": nombre_hoja,
                        "Prueba": prueba.numero,
                        "Alcance": prueba.alcance,
                        "Mecanismo": prueba.mecanismo,
                        "Mejor k global": None,
                        "Mejor perdida global": None,
                    }
                )

        wb.save(ruta_salida)
        print(f"  Hoja {nombre_hoja} guardada.")

    wb.save(ruta_salida)

    resumen_path = ruta_salida.with_name(ruta_salida.stem + "_resumen_qnodes.xlsx")
    if resumen_filas:
        pd.DataFrame(resumen_filas).to_excel(resumen_path, index=False)

    print(f"\nResultados QNodes en plantilla: {ruta_salida}")
    print(f"Resumen detallado:              {resumen_path}")


def iniciar():
    """Punto de entrada en modo lote sobre DatosPruebas2026_1.xlsx."""
    ruta_entrada = Path(os.getenv("QNODES_INPUT_XLSX", str(DEFAULT_EXCEL)))
    if not ruta_entrada.exists():
        print(f"No se encontro el Excel de entrada: {ruta_entrada}")
        print("Defina QNODES_INPUT_XLSX con la ruta a DatosPruebas2026_1.xlsx")
        return

    ruta_salida = Path(
        os.getenv(
            "QNODES_OUTPUT_XLSX",
            str(ruta_entrada.with_name(ruta_entrada.stem + "_resultados.xlsx")),
        )
    )
    cantidad = os.getenv("QNODES_CANTIDAD", os.getenv("GEOMIP_CANTIDAD"))
    cantidad = int(cantidad) if cantidad else None
    hoja = os.getenv("QNODES_HOJA", os.getenv("GEOMIP_HOJA"))

    ejecutar_datos_pruebas2026(
        ruta_entrada, ruta_salida, cantidad=cantidad, hoja_filtro=hoja
    )
