"""Lector/escritor para DatosPruebas2026_1.xlsx (formato del curso)."""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

import pandas as pd
from openpyxl import load_workbook

from src.funcs.base import get_labels

LETRAS = get_labels(40)

# Columnas Geometric en la plantilla (0-indexed): Particion, Perdida, Tiempo
COLUMNAS_GEOMETRIC = {
    2: (6, 7, 8),
    3: (12, 13, 14),
    4: (18, 19, 20),
    5: (24, 25, 26),
}

FILA_ESTADO = 0
FILA_SISTEMA = 1
FILA_CANDIDATO = 2
FILA_DATOS_INICIO = 5


@dataclass
class Prueba2026:
    numero: int
    fila_excel: int
    estado_inicial: str
    sistema: str
    candidato: str
    condiciones: str
    alcance: str
    mecanismo: str
    n_nodos: int
    pagina_tpm: str
    hoja: str


def _celda_valor(df: pd.DataFrame, fila: int, col: int) -> str:
    if fila >= len(df) or col >= len(df.columns):
        return ""
    val = df.iloc[fila, col]
    if pd.isna(val):
        return ""
    return str(val).strip()


def letras_a_binario(texto: str, n_bits: int) -> str:
    if not texto:
        return "0" * n_bits
    texto = str(texto).strip().upper()
    if all(c in "01" for c in texto):
        if len(texto) == n_bits:
            return texto
        if len(texto) < n_bits:
            return texto.ljust(n_bits, "0")
    binario = ["0"] * n_bits
    for letra in texto:
        if letra in LETRAS[:n_bits]:
            binario[LETRAS.index(letra)] = "1"
    return "".join(binario)


def sistema_candidato_a_condicion(sistema: str, candidato: str, n_bits: int) -> str:
    """Bit 1 = variable en candidato; bit 0 = condicionada (background)."""
    cand_set = set(str(candidato).strip().upper())
    bits = []
    for i in range(n_bits):
        letra = LETRAS[i]
        bits.append("1" if letra in cand_set else "0")
    return "".join(bits)


def parsear_nombre_hoja(nombre: str) -> tuple[int, str] | None:
    m = re.match(r"(\d+)([A-Z])-Elementos\s*$", nombre.strip(), re.I)
    if not m:
        return None
    return int(m.group(1)), m.group(2).upper()


def leer_hoja_pruebas(ruta: Path, nombre_hoja: str) -> list[Prueba2026]:
    df = pd.read_excel(ruta, sheet_name=nombre_hoja, header=None)
    parsed = parsear_nombre_hoja(nombre_hoja)
    if parsed is None:
        return []
    n_nodos, pagina = parsed

    estado = _celda_valor(df, FILA_ESTADO, 1)
    if not estado or not all(c in "01" for c in estado):
        estado = "1" + "0" * (n_nodos - 1)
    n_bits = len(estado)

    sistema = _celda_valor(df, FILA_SISTEMA, 1)
    candidato = _celda_valor(df, FILA_CANDIDATO, 1)
    condiciones = sistema_candidato_a_condicion(sistema, candidato, n_bits)

    pruebas: list[Prueba2026] = []
    for fila in range(FILA_DATOS_INICIO, len(df)):
        num_raw = _celda_valor(df, fila, 0)
        if not num_raw or num_raw.lower().startswith("#"):
            continue
        try:
            numero = int(float(num_raw))
        except ValueError:
            continue

        alcance_txt = _celda_valor(df, fila, 1)
        mecanismo_txt = _celda_valor(df, fila, 2)
        if not alcance_txt and not mecanismo_txt:
            continue

        pruebas.append(
            Prueba2026(
                numero=numero,
                fila_excel=fila,
                estado_inicial=estado,
                sistema=sistema,
                candidato=candidato,
                condiciones=condiciones,
                alcance=letras_a_binario(alcance_txt, n_bits),
                mecanismo=letras_a_binario(mecanismo_txt, n_bits),
                n_nodos=n_nodos,
                pagina_tpm=pagina,
                hoja=nombre_hoja,
            )
        )
    return pruebas


def iter_hojas_pruebas(ruta: Path) -> Iterator[tuple[str, list[Prueba2026]]]:
    xl = pd.ExcelFile(ruta)
    for nombre in xl.sheet_names:
        if parsear_nombre_hoja(nombre):
            pruebas = leer_hoja_pruebas(ruta, nombre)
            if pruebas:
                yield nombre, pruebas


def escribir_resultado_fila(
    ws,
    fila: int,
    resultados_por_k: dict,
    mejor_k: int | None,
    mejor_perdida: float | None,
) -> None:
    """Escribe columnas Geometric (k=2..5) en la fila de la prueba."""
    for k, (col_part, col_perd, col_time) in COLUMNAS_GEOMETRIC.items():
        r = resultados_por_k.get(k)
        if not r:
            continue
        ws.cell(row=fila + 1, column=col_part + 1, value=r.get("particion"))
        perd = r.get("perdida")
        if perd is not None:
            ws.cell(row=fila + 1, column=col_perd + 1, value=perd)
        tiempo = r.get("tiempo")
        if tiempo is not None:
            ws.cell(row=fila + 1, column=col_time + 1, value=tiempo)

    if mejor_k is not None:
        ws.cell(row=fila + 1, column=28, value=f"MEJOR k={mejor_k}")
        if mejor_perdida is not None:
            ws.cell(row=fila + 1, column=29, value=mejor_perdida)


def guardar_workbook(ruta_entrada: Path, ruta_salida: Path) -> None:
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    import shutil
    if ruta_salida.resolve() != ruta_entrada.resolve():
        shutil.copy2(ruta_entrada, ruta_salida)