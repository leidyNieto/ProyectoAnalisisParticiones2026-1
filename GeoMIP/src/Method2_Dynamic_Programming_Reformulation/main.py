from src.controllers.manager import Manager
from src.controllers.strategies.geometric import GeometricSIA
try:
    from src.controllers.strategies.kpartition import KPartitionSIA
except ImportError:
    KPartitionSIA = None
from src.funcs.excel_pruebas2026 import (
    escribir_resultado_fila,
    iter_hojas_pruebas,
    parsear_nombre_hoja,
)
import multiprocessing
import numpy as np
import pandas as pd
import os
import re
import shutil
import traceback
from pathlib import Path
from openpyxl import load_workbook

METHOD2_ROOT = Path(__file__).resolve().parents[1]
GEOMIP_ROOT = Path(__file__).resolve().parents[3]

LETRAS = "ABCDEFGHIJKLMNOPQRST"
DEFAULT_EXCEL = GEOMIP_ROOT / "results" / "DatosPruebas2026_1.xlsx"


def convertir_a_binario(texto: str, n_bits: int) -> str:
    if texto is None or (isinstance(texto, float) and np.isnan(texto)):
        return "1" * n_bits
    texto = str(texto).strip().upper()
    if all(c in "01" for c in texto):
        if len(texto) == n_bits:
            return texto
        if len(texto) < n_bits:
            return texto.ljust(n_bits, "0")
    binario = ["0"] * n_bits
    for letra in texto:
        if letra in LETRAS[:n_bits]:
            binario[LETRAS.index(letra)] = "1"
    return "".join(binario)


def resolver_tpm_path(estado_inicio: str, pagina: str = "A") -> Path:
    n = len(estado_inicio)
    candidatos_letra = [pagina] + [c for c in "ABCDE" if c != pagina]
    bases = (
        METHOD2_ROOT / "src" / ".samples",
        METHOD2_ROOT / ".samples",
        GEOMIP_ROOT / "data" / "samples",
    )
    for letra in candidatos_letra:
        sample_name = f"N{n}{letra}.csv"
        for base in bases:
            path = base / sample_name
            if path.exists():
                return path
    disponibles = []
    for base in bases:
        if base.exists():
            disponibles.extend(p.name for p in base.glob(f"N{n}*.csv"))
    extra = f" Disponibles para N{n}: {', '.join(sorted(set(disponibles))) or 'ninguno'}."
    raise FileNotFoundError(
        f"No se encontro TPM para N{n} (pagina {pagina}).{extra}"
    )


def asegurar_tpm(n_nodos: int, pagina: str) -> Path:
    """
    Busca la TPM; si no existe la genera con Manager.generar_red
    (misma logica que usar manager.py manualmente).
    """
    estado = "1" + "0" * (n_nodos - 1)
    auto_gen = os.getenv("GEOMIP_AUTO_TPM", "1").lower() in ("1", "true", "si", "s")

    try:
        return resolver_tpm_path(estado, pagina)
    except FileNotFoundError:
        pass

    print(f"\n>>> TPM N{n_nodos}{pagina} no encontrada. Generando con Manager.generar_red ...")
    gestor = Manager(estado_inicial=estado)
    nombre = gestor.generar_red(
        dimensiones=n_nodos,
        pagina=pagina,
        datos_discretos=False,
        auto_confirm=auto_gen,
    )
    if not nombre:
        raise FileNotFoundError(
            f"No se pudo generar N{n_nodos}{pagina}.csv. "
            "Para redes >1 GB use GEOMIP_AUTO_TPM=1."
        )
    path = gestor.ruta_base / nombre
    if not path.exists():
        raise FileNotFoundError(f"Generacion reporto {nombre} pero no existe en {path}")
    print(f">>> TPM generada: {path}\n")
    return path


def _estrategia_desde_env():
    nombre = os.getenv("GEOMIP_STRATEGY", "Geometric").strip().lower()
    if nombre in ("kpartition", "bruteforce", "fuerza_bruta", "fb"):
        if KPartitionSIA is None:
            raise ImportError("KPartitionSIA no disponible; use Geometric.")
        return KPartitionSIA
    return GeometricSIA


def _fmt_num(val):
    if val is None:
        return None
    return str(val).replace(".", ",")


def _resolver_prueba(
    estado, condiciones, alcance, mecanismo, tpm, estrategia_cls
) -> dict:
    gestor = Manager(estado_inicial=estado)
    sia = estrategia_cls(gestor)
    sol = sia.aplicar_estrategia(condiciones, alcance, mecanismo, tpm)
    por_k = getattr(sol, "resultados_por_k", {}) or {}
    return {
        "particion": sol.particion,
        "perdida": sol.perdida,
        "tiempo": sol.tiempo_ejecucion,
        "k": sol.k_particiones,
        "por_k": por_k,
    }


def _worker_prueba(args, queue):
    try:
        queue.put(_resolver_prueba(*args))
    except Exception:
        traceback.print_exc()
        queue.put(None)


def _fila_resumen(resultado: dict) -> dict:
    por_k = resultado.get("por_k") or {}
    fila = {
        "Mejor k global": resultado.get("k"),
        "Mejor particion global": resultado.get("particion"),
        "Mejor perdida global": _fmt_num(resultado.get("perdida")),
        "Tiempo total (s)": _fmt_num(resultado.get("tiempo")),
    }
    for k in range(2, 6):
        r = por_k.get(k, {})
        fila[f"k{k} Particion"] = r.get("particion")
        fila[f"k{k} Perdida"] = _fmt_num(r.get("perdida"))
        fila[f"k{k} Tiempo (s)"] = _fmt_num(r.get("tiempo"))
    return fila


def ejecutar_datos_pruebas2026(
    ruta_entrada: Path,
    ruta_salida: Path | None = None,
    cantidad: int | None = None,
    hoja_filtro: str | None = None,
):
    """
    Lee DatosPruebas2026_1.xlsx, ejecuta Geometric para cada prueba,
    escribe resultados por k (2..5) en columnas Geometric y el mejor global.
    """
    ruta_salida = ruta_salida or ruta_entrada.with_name(
        ruta_entrada.stem + "_resultados.xlsx"
    )
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    if ruta_salida.resolve() != ruta_entrada.resolve():
        shutil.copy2(ruta_entrada, ruta_salida)

    wb = load_workbook(ruta_salida)
    estrategia_cls = _estrategia_desde_env()
    resumen_filas = []
    timeout = int(os.getenv("GEOMIP_TIMEOUT", "3600"))

    for nombre_hoja, pruebas in iter_hojas_pruebas(ruta_entrada):
        if hoja_filtro and hoja_filtro.lower() not in nombre_hoja.lower():
            continue
        if nombre_hoja not in wb.sheetnames:
            continue

        ws = wb[nombre_hoja]
        parsed = parsear_nombre_hoja(nombre_hoja)
        if not parsed:
            continue
        n_nodos, pagina = parsed

        try:
            tpm_path = asegurar_tpm(n_nodos, pagina)
        except FileNotFoundError as exc:
            print(f"\n=== OMITIDA hoja {nombre_hoja}: {exc} ===")
            continue

        # Cargar TPM: genfromtxt puede lanzar MemoryError en archivos muy grandes (p.ej. N=25).
        try:
            tpm = np.genfromtxt(tpm_path, delimiter=",")
        except MemoryError:
            print(f"\n=== TPM demasiado grande para cargar con numpy: {tpm_path} ===")
            print("Intentando fallback con pandas usando dtype=float32 (menos memoria). Si falla, use TPM menor o aumente RAM.")
            try:
                import pandas as _pd

                tpm = _pd.read_csv(tpm_path, header=None, dtype=_pd.Float32Dtype() if hasattr(_pd, 'Float32Dtype') else 'float32').values
            except Exception as exc:
                print(f"Error al cargar TPM con pandas: {exc}")
                print(f"Omitiendo hoja {nombre_hoja}.")
                continue
        print(f"\n=== Hoja: {nombre_hoja} | TPM: {tpm_path.name} ===")

        limite = cantidad if cantidad else len(pruebas)
        for prueba in pruebas[:limite]:
            print(
                f"  Prueba #{prueba.numero} | alcance={prueba.alcance} "
                f"mecanismo={prueba.mecanismo}"
            )

            queue = multiprocessing.Queue()
            args = (
                prueba.estado_inicial,
                prueba.condiciones,
                prueba.alcance,
                prueba.mecanismo,
                tpm,
                estrategia_cls,
            )
            proc = multiprocessing.Process(target=_worker_prueba, args=(args, queue))
            proc.start()
            proc.join(timeout=timeout)

            if proc.is_alive():
                proc.terminate()
                proc.join()
                resultado = None
                print(f"  Prueba #{prueba.numero} - TIMEOUT")
            else:
                resultado = queue.get() if not queue.empty() else None

            if resultado:
                escribir_resultado_fila(
                    ws,
                    prueba.fila_excel,
                    resultado["por_k"],
                    resultado["k"],
                    resultado["perdida"],
                )
                print(
                    f"  -> Mejor k={resultado['k']} perdida={resultado['perdida']:.4f} "
                    f"tiempo={resultado['tiempo']:.3f}s"
                )
                # Mostrar la partición completa (igual que se escribe en el Excel)
                if resultado.get("particion"):
                    print(f"  -> Particion mejor (formato):\n{resultado.get('particion')}")
                # Mostrar particiones y métricas por k
                for k, r in sorted(resultado["por_k"].items()):
                    part_text = r.get("particion") or ""
                    perd = r.get("perdida")
                    tiempo = r.get("tiempo")
                    print(
                        f"     k={k}: particion={part_text} perdida={perd:.4f} tiempo={tiempo:.3f}s"
                        if perd is not None and tiempo is not None
                        else f"     k={k}: particion={part_text}"
                    )
                resumen_filas.append({
                    "Hoja": nombre_hoja,
                    "Prueba": prueba.numero,
                    "Alcance": prueba.alcance,
                    "Mecanismo": prueba.mecanismo,
                    **_fila_resumen(resultado),
                })
            else:
                resumen_filas.append({
                    "Hoja": nombre_hoja,
                    "Prueba": prueba.numero,
                    "Alcance": prueba.alcance,
                    "Mecanismo": prueba.mecanismo,
                    "Mejor k global": None,
                    "Mejor perdida global": None,
                })

        wb.save(ruta_salida)
        print(f"  Hoja {nombre_hoja} guardada.")

    wb.save(ruta_salida)

    resumen_path = ruta_salida.with_name(ruta_salida.stem + "_resumen.xlsx")
    if resumen_filas:
        pd.DataFrame(resumen_filas).to_excel(resumen_path, index=False)

    print(f"\nResultados en plantilla: {ruta_salida}")
    print(f"Resumen detallado:       {resumen_path}")


def iniciar():
    ruta_entrada = Path(
        os.getenv("GEOMIP_INPUT_XLSX", str(DEFAULT_EXCEL))
    )
    if not ruta_entrada.exists():
        ruta_entrada = Path(
            os.getenv("GEOMIP_INPUT_XLSX", str(GEOMIP_ROOT / "results" / "DatosPruebas2026_1.xlsx"))
        )

    ruta_salida = Path(
        os.getenv(
            "GEOMIP_OUTPUT_XLSX",
            str(ruta_entrada.with_name(ruta_entrada.stem + "_resultados.xlsx")),
        )
    )
    cantidad = os.getenv("GEOMIP_CANTIDAD")
    cantidad = int(cantidad) if cantidad else None
    hoja = os.getenv("GEOMIP_HOJA")

    if "datospruebas2026" in ruta_entrada.name.lower() or os.getenv("GEOMIP_FORMATO", "2026") == "2026":
        ejecutar_datos_pruebas2026(
            ruta_entrada, ruta_salida, cantidad=cantidad, hoja_filtro=hoja
        )
    else:
        print(f"Formato Excel no reconocido: {ruta_entrada}")
        print("Use DatosPruebas2026_1.xlsx o GEOMIP_FORMATO=2026")
