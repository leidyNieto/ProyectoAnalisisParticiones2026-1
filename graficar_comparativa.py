#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
graficar_comparativa.py
=======================
Lee el Excel de resultados (QNodes vs Geometric, una hoja por subsistema con
bloques [Particion, Perdida, Tiempo] por estrategia y por k) y genera las
gráficas comparativas + una tabla resumen.

Produce:
  1. fig1_tiempo.png          -> Tiempo de ejecución por estrategia vs tamaño (asc), por k.
  2. fig2_perdida.png         -> Pérdida (EMD) por estrategia y tamaño, por k.
  3. fig3_variacion_perdida.png -> Variación de pérdida vs QNodes (absoluta + relativa), barras.
  4. fig4_particion.png       -> % de pruebas con la misma k-partición, barras.
  4. resumen_comparativo.csv / .xlsx -> Tabla agregada por (tamaño, k).

Uso:
  python graficar_comparativa.py                       # usa rutas por defecto
  python graficar_comparativa.py -i resultados.xlsx -o ./graficas
  python graficar_comparativa.py --ks 2 3 4 5

Requisitos:  pip install pandas openpyxl matplotlib
"""
from __future__ import annotations
import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter

# ----------------------------------------------------------------------------
# CONFIGURACIÓN (edítala si cambian nombres/colores)
# ----------------------------------------------------------------------------
DEFAULT_INPUT = "results/DatosPruebas2026_1_resultados.xlsx"
DEFAULT_OUTDIR = "."
COLOR_Q = "#D7642E"   # QNodes  (naranja)
COLOR_G = "#2E6FB0"   # Geometric (azul)
K_COLORS = {2: "#1b9e77", 3: "#7570b3", 4: "#e7298a", 5: "#d95f02"}
# Etiquetas tal como aparecen en la fila de cabecera del Excel:
LABEL_Q = "QNodes"
LABEL_G = "Geometric"

plt.rcParams.update({
    "font.size": 11, "axes.grid": True, "grid.alpha": .3, "axes.edgecolor": "#888",
    "figure.dpi": 130, "savefig.bbox": "tight", "font.family": "DejaVu Sans",
})


# ----------------------------------------------------------------------------
# 1) LECTURA Y PARSEO DEL EXCEL
# ----------------------------------------------------------------------------
def _size_from_sheet(name: str):
    """'10A-Elementos' -> 10 ; devuelve None si la hoja no es de resultados."""
    m = re.match(r"\s*(\d+)", name.strip())
    return int(m.group(1)) if m else None


def _find_blocks(rows):
    """
    Detecta automáticamente la estructura de columnas a partir de las cabeceras.
    Busca la fila que contiene las etiquetas de estrategia (QNodes/Geometric) y
    la fila inmediatamente inferior con 'Partición/Pérdida/Tiempo'.

    Devuelve:
        header_data_start: índice (0-based) de la primera fila de datos
        kmap: dict {k: {'q': (cp, cl, ct), 'g': (cp, cl, ct)}}
              donde cp/cl/ct son índices de columna (0-based) de
              Partición, Pérdida y Tiempo.
    """
    strat_row_idx = None
    for i, r in enumerate(rows[:12]):
        cells = [str(c) if c is not None else "" for c in r]
        if any(LABEL_Q in c for c in cells) and any(LABEL_G in c for c in cells):
            strat_row_idx = i
            break
    if strat_row_idx is None:
        raise ValueError("No se encontró la fila de cabecera con QNodes/Geometric.")

    strat_row = rows[strat_row_idx]
    # Cada etiqueta de estrategia abre un bloque de 3 columnas: Partición, Pérdida, Tiempo
    blocks = []  # lista de ('q'|'g', col_inicio)
    for col, c in enumerate(strat_row):
        if c is None:
            continue
        s = str(c)
        if LABEL_Q in s:
            blocks.append(("q", col))
        elif LABEL_G in s:
            blocks.append(("g", col))

    # Empareja bloques consecutivos q,g -> una k (empezando en k=2)
    kmap = {}
    k = 2
    bi = 0
    while bi + 1 < len(blocks) + 1 and bi < len(blocks):
        strat, col = blocks[bi]
        cols = (col, col + 1, col + 2)  # Partición, Pérdida, Tiempo
        # ¿el siguiente bloque es de la otra estrategia? -> misma k
        if strat == "q" and bi + 1 < len(blocks) and blocks[bi + 1][0] == "g":
            gcol = blocks[bi + 1][1]
            kmap[k] = {"q": cols, "g": (gcol, gcol + 1, gcol + 2)}
            k += 1
            bi += 2
        else:
            # bloque suelto: lo asignamos igualmente
            kmap.setdefault(k, {})[strat] = cols
            bi += 1
    data_start = strat_row_idx + 2  # fila estrategia + fila sub-cabecera
    return data_start, kmap


def _canon(s):
    """Representación canónica de una partición = conjunto de grupos de elementos
    (mayúsculas, t+1) ignorando orden. Sirve para comparar si dos estrategias
    hallan la MISMA k-partición aunque el texto difiera de formato."""
    if not s or not isinstance(s, str):
        return None
    up_groups = re.findall(r"[A-Z](?:,[A-Z])*", s.replace("\n", " "))
    if not up_groups:
        return None
    return frozenset(frozenset(g.split(",")) for g in up_groups)


def cargar_resultados(path: str, ks_filtro=None) -> pd.DataFrame:
    """Lee el Excel y devuelve un DataFrame 'tidy' con una fila por (hoja, prueba, k)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    registros = []
    for sh in wb.sheetnames:
        size = _size_from_sheet(sh)
        if size is None:
            continue  # hojas como 'plataformas', 'Requerimientos', etc.
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        try:
            data_start, kmap = _find_blocks(rows)
        except ValueError:
            continue
        if ks_filtro:
            kmap = {k: v for k, v in kmap.items() if k in ks_filtro}

        prueba = 0
        for r in rows[data_start:]:
            # fila válida = tiene pérdida en algún bloque
            tiene = any(
                (("q" in b and r[b["q"][1]] is not None) or
                 ("g" in b and r[b["g"][1]] is not None))
                for b in kmap.values()
            )
            if not tiene:
                continue
            prueba += 1
            alcance = r[1] if len(r) > 1 else None
            mecanismo = r[2] if len(r) > 2 else None
            for k, b in kmap.items():
                rec = dict(size=size, sheet=sh.strip(), prueba=prueba, k=k,
                           alcance=alcance, mecanismo=mecanismo,
                           q_part=None, q_loss=None, q_time=None,
                           g_part=None, g_loss=None, g_time=None)
                if "q" in b:
                    cp, cl, ct = b["q"]
                    rec.update(q_part=r[cp], q_loss=r[cl], q_time=r[ct])
                if "g" in b:
                    cp, cl, ct = b["g"]
                    rec.update(g_part=r[cp], g_loss=r[cl], g_time=r[ct])
                registros.append(rec)
    wb.close()

    df = pd.DataFrame(registros)
    for c in ["q_loss", "q_time", "g_loss", "g_time"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    # ¿misma k-partición? (solo donde ambas estrategias tienen partición)
    def _same(row):
        a, b = _canon(row["q_part"]), _canon(row["g_part"])
        return None if (a is None or b is None) else (a == b)
    df["same_part"] = df.apply(_same, axis=1)
    return df.sort_values(["size", "k", "prueba"]).reset_index(drop=True)


def construir_resumen(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega por (tamaño, k): tiempos/pérdidas medias, speedup, % misma partición, Δpérdida."""
    summ = df.groupby(["size", "k"]).agg(
        n=("prueba", "count"),
        q_time=("q_time", "mean"), g_time=("g_time", "mean"),
        q_loss=("q_loss", "mean"), g_loss=("g_loss", "mean"),
    ).reset_index()
    tmp = df.dropna(subset=["same_part"]).copy()
    # same_part puede venir como object (True/False/None); forzar a numérico
    tmp["same_part"] = tmp["same_part"].astype(bool).astype(float)
    ag = tmp.groupby(["size", "k"])["same_part"].agg(["sum", "count"]).reset_index()
    ag["sum"] = ag["sum"].astype(float)
    ag["count"] = ag["count"].astype(float)
    ag["pct_same"] = (100 * ag["sum"] / ag["count"]).round(1)
    summ = summ.merge(ag[["size", "k", "pct_same"]], on=["size", "k"], how="left")
    summ["speedup"] = (summ.q_time / summ.g_time).round(2)
    summ["dloss_mean"] = summ.g_loss - summ.q_loss
    return summ


# ----------------------------------------------------------------------------
# 2) GRÁFICAS
# ----------------------------------------------------------------------------
def _bars_grouped(ax, cats, series, colors, width_total=0.8, logy=False,
                  errors=None, fmt="{:.2g}"):
    """Dibuja barras agrupadas. `series` = dict {etiqueta: lista de valores alineada a `cats`}.
    Omite barras con valor None/NaN. Anota el valor encima de cada barra."""
    import numpy as np
    labels = list(series.keys())
    n = len(labels)
    w = width_total / max(n, 1)
    x = np.arange(len(cats), dtype=float)
    for i, lab in enumerate(labels):
        vals = np.array([np.nan if v is None else v for v in series[lab]], dtype=float)
        off = (i - (n - 1) / 2) * w
        err = None
        if errors is not None:
            err = np.array([np.nan if v is None else v for v in errors[lab]], dtype=float)
        bars = ax.bar(x + off, np.nan_to_num(vals, nan=0.0), w, color=colors[i],
                      edgecolor="white", label=lab,
                      yerr=err, capsize=3, error_kw=dict(lw=1, ecolor="#555"))
        for b, v in zip(bars, vals):
            if np.isnan(v):
                b.set_height(0); continue
            y = v
            ax.annotate(fmt.format(v), (b.get_x() + b.get_width() / 2, y),
                        textcoords="offset points", xytext=(0, 3),
                        ha="center", fontsize=7.5, rotation=0)
    ax.set_xticks(x); ax.set_xticklabels([str(c) for c in cats])
    if logy:
        ax.set_yscale("log")


def fig_tiempo(df, summ, ks, sizes, outdir):
    """Barras agrupadas: tiempo medio (± std) por estrategia y tamaño, un panel por k.
    Ejes Y compartidos y escala log -> fácil de comparar entre paneles."""
    stats = (df.groupby(["size", "k"])
               .agg(qm=("q_time", "mean"), qs=("q_time", "std"),
                    gm=("g_time", "mean"), gs=("g_time", "std")).reset_index())
    fig, axes = plt.subplots(1, len(ks), figsize=(3.7 * len(ks), 4.4), sharey=True)
    if len(ks) == 1:
        axes = [axes]
    for ax, k in zip(axes, ks):
        d = stats[stats.k == k].set_index("size").reindex(sizes)
        series = {"QNodes": d["qm"].tolist(), "Geometric": d["gm"].tolist()}
        errors = {"QNodes": d["qs"].tolist(), "Geometric": d["gs"].tolist()}
        _bars_grouped(ax, sizes, series, [COLOR_Q, COLOR_G], logy=True,
                      errors=errors, fmt="{:.3g}")
        ax.set_title(f"k = {k}", fontweight="bold")
        ax.set_xlabel("Tamaño (nº elementos)")
        ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v:g}"))
    axes[0].set_ylabel("Tiempo de ejecución medio (s, escala log)")
    axes[0].legend(loc="upper left", framealpha=.95)
    fig.suptitle("Tiempo de ejecución por estrategia y tamaño de subsistema (orden ascendente)",
                 fontweight="bold", y=1.03, fontsize=13)
    out = outdir / "fig1_tiempo.png"
    fig.savefig(out); plt.close(fig)
    return out


def fig_particion_perdida(df, summ, ks, outdir):
    """Barras agrupadas. Izq: pérdida media (EMD) por estrategia y tamaño, un grupo por k.
    Der: % de pruebas en que Geometric halla la MISMA k-partición que QNodes."""
    sizes_both = sorted(df[df.g_loss.notna()]["size"].unique())
    fig, axes = plt.subplots(1, len(ks), figsize=(3.7 * len(ks), 4.4), sharey=True)
    if len(ks) == 1:
        axes = [axes]
    stats = (df.groupby(["size", "k"])
               .agg(qm=("q_loss", "mean"), gm=("g_loss", "mean")).reset_index())
    allsizes = sorted(df["size"].unique())
    for ax, k in zip(axes, ks):
        d = stats[stats.k == k].set_index("size").reindex(allsizes)
        series = {"QNodes": d["qm"].tolist(), "Geometric": d["gm"].tolist()}
        _bars_grouped(ax, allsizes, series, [COLOR_Q, COLOR_G], fmt="{:.3f}")
        ax.set_title(f"k = {k}", fontweight="bold")
        ax.set_xlabel("Tamaño (nº elementos)")
    axes[0].set_ylabel("Pérdida media (EMD)")
    axes[0].legend(loc="upper left", framealpha=.95)
    fig.suptitle("Pérdida (EMD) por estrategia y tamaño de subsistema, por k",
                 fontweight="bold", y=1.03, fontsize=13)
    out = outdir / "fig2_perdida.png"
    fig.savefig(out); plt.close(fig)
    return out


def fig_variacion_perdida(df, ks, outdir):
    """Dos paneles de BARRAS, ambos comparan la pérdida (EMD) de Geometric frente a QNodes
    (QNodes = referencia), por k y por tamaño N.
      Izq: desviación ABSOLUTA media  = media( EMD_Geometric − EMD_QNodes )  [unidades de EMD]
      Der: desviación RELATIVA media  = media( |Δ| / EMD_QNodes ) · 100      [%]
    En ambos, 0 = Geometric coincide con QNodes; valores > 0 = Geometric pierde más."""
    both = df[df.g_loss.notna() & df.q_loss.notna()].copy()
    both["dloss"] = both.g_loss - both.q_loss
    both["relpct"] = np.where(both.q_loss > 1e-9,
                              100 * both.dloss.abs() / both.q_loss, 0.0)
    sizes_both = sorted(both["size"].unique())
    palette = [COLOR_Q, COLOR_G, "#1b9e77", "#7570b3"]

    fig, (axA, axB) = plt.subplots(1, 2, figsize=(13, 4.8))

    # Izq: desviación absoluta media (EMD)
    absm = both.groupby(["size", "k"])["dloss"].mean().reset_index()
    seriesA = {f"N={s}": absm[absm["size"] == s].set_index("k").reindex(ks)["dloss"].tolist()
               for s in sizes_both}
    _bars_grouped(axA, ks, seriesA, palette[:len(sizes_both)], fmt="{:.4f}")
    axA.axhline(0, color="#c00", lw=1.4, label="Referencia QNodes (Δ=0)")
    axA.set_xlabel("k (nº de partes)")
    axA.set_ylabel("Δ pérdida media = EMD(Geometric) − EMD(QNodes)")
    axA.set_title("Desviación ABSOLUTA media (unidades EMD)", fontweight="bold", fontsize=11)
    axA.legend(loc="upper left", fontsize=9)

    # Der: desviación relativa media (%)
    rel = both.groupby(["size", "k"])["relpct"].mean().reset_index()
    seriesB = {f"N={s}": rel[rel["size"] == s].set_index("k").reindex(ks)["relpct"].tolist()
               for s in sizes_both}
    _bars_grouped(axB, ks, seriesB, palette[:len(sizes_both)], fmt="{:.2f}")
    axB.axhline(0, color="#c00", lw=1.4, label="Referencia QNodes (0%)")
    axB.set_xlabel("k (nº de partes)")
    axB.set_ylabel("Error relativo medio |Δ| / QNodes (%)")
    axB.set_title("Desviación RELATIVA media (%)", fontweight="bold", fontsize=11)
    axB.legend(loc="upper left", fontsize=9)

    fig.suptitle("Variación de pérdida (EMD): Geometric comparado contra QNodes (referencia)",
                 fontweight="bold", y=1.03, fontsize=13)
    out = outdir / "fig3_variacion_perdida.png"
    fig.savefig(out); plt.close(fig)
    return out


def fig_particion(df, ks, outdir):
    """Barras: % de pruebas en que Geometric halla la MISMA k-partición que QNodes, por N."""
    same = df.dropna(subset=["same_part"]).copy()
    same["same_part"] = same["same_part"].astype(bool).astype(float)
    pct = (same.groupby(["size", "k"])["same_part"].mean() * 100).reset_index()
    sizes_both = sorted(same["size"].unique())
    palette = [COLOR_Q, COLOR_G, "#1b9e77", "#7570b3"]

    fig, ax = plt.subplots(figsize=(7, 4.6))
    series = {f"N={s}": pct[pct["size"] == s].set_index("k").reindex(ks)["same_part"].tolist()
              for s in sizes_both}
    _bars_grouped(ax, ks, series, palette[:len(sizes_both)], fmt="{:.0f}")
    ax.set_ylim(0, 112); ax.axhline(100, ls=":", color="#999")
    ax.set_xlabel("k (nº de partes)")
    ax.set_ylabel("% de pruebas con la MISMA k-partición")
    ax.set_title("¿Geometric halla la misma k-partición que QNodes?", fontweight="bold")
    ax.legend()
    out = outdir / "fig4_particion.png"
    fig.savefig(out); plt.close(fig)
    return out


def exportar_resumen(summ: pd.DataFrame, outdir: Path):
    out = summ.copy()
    out.columns = ["Tamaño", "k", "Nº pruebas", "Tiempo QNodes (s)", "Tiempo Geometric (s)",
                   "Pérdida QNodes", "Pérdida Geometric", "% misma k-part.",
                   "Speedup (Q/G)", "Δpérdida media"]
    out = out.round({"Tiempo QNodes (s)": 3, "Tiempo Geometric (s)": 3, "Pérdida QNodes": 5,
                     "Pérdida Geometric": 5, "Speedup (Q/G)": 2, "Δpérdida media": 6})
    out.to_csv(outdir / "resumen_comparativo.csv", index=False)
    try:
        with pd.ExcelWriter(outdir / "resumen_comparativo.xlsx") as xw:
            out.to_excel(xw, index=False, sheet_name="Resumen")
    except Exception as e:  # openpyxl ausente, etc.
        print(f"[aviso] no se pudo escribir xlsx: {e}")
    return out


# ----------------------------------------------------------------------------
# 3) MAIN
# ----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Genera gráficas comparativas QNodes vs Geometric.")
    ap.add_argument("-i", "--input", default=DEFAULT_INPUT, help="Excel de resultados.")
    ap.add_argument("-o", "--outdir", default=DEFAULT_OUTDIR, help="Carpeta de salida.")
    ap.add_argument("--ks", type=int, nargs="+", default=None,
                    help="k a incluir (p.ej. --ks 2 3 4 5). Por defecto: las que haya.")
    args = ap.parse_args()

    inp = Path(args.input)
    outdir = Path(args.outdir); outdir.mkdir(parents=True, exist_ok=True)
    if not inp.exists():
        raise SystemExit(f"No existe el archivo: {inp}")

    print(f"[1/4] Leyendo {inp} ...")
    df = cargar_resultados(str(inp), ks_filtro=set(args.ks) if args.ks else None)
    if df.empty:
        raise SystemExit("No se encontraron datos de resultados en el Excel.")
    ks = sorted(df["k"].unique())
    sizes = sorted(df["size"].unique())
    print(f"      Tamaños: {sizes} | k: {ks} | filas: {len(df)}")

    print("[2/4] Resumen agregado ...")
    summ = construir_resumen(df)
    exportar_resumen(summ, outdir)

    print("[3/4] Generando gráficas ...")
    f1 = fig_tiempo(df, summ, ks, sizes, outdir)
    f2 = fig_particion_perdida(df, summ, ks, outdir)
    f3 = fig_variacion_perdida(df, ks, outdir)
    f4 = fig_particion(df, ks, outdir)

    print("[4/4] Listo. Archivos generados:")
    for p in [f1, f2, f3, f4, outdir / "resumen_comparativo.csv", outdir / "resumen_comparativo.xlsx"]:
        print("   -", p)


if __name__ == "__main__":
    main()